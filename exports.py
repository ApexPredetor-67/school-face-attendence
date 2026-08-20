"""Attendance export helpers.

This module is intentionally kept local to the project so ``from exports``
resolves to this file rather than a third-party package named ``exports``.
"""
from io import BytesIO
from typing import Iterable, Mapping

COLUMNS = [
    ("admission_number", "Admission No"),
    ("roll_number", "Roll No"),
    ("name", "Name"),
    ("class_section", "Class"),
    ("attendance", "Attendance"),
    ("time_in", "Time"),
]


def _cell_value(value):
    """Return a value that Excel/reportlab can safely render."""
    if value is None:
        return ""
    return str(value)


def build_xlsx(rows: Iterable[Mapping], title: str = "Attendance Register", subtitle: str = "") -> BytesIO:
    """Build an in-memory .xlsx attendance register."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.freeze_panes = "A4" if subtitle else "A3"
    ws.auto_filter.ref = "A3:F3" if subtitle else "A2:F2"

    ws.merge_cells("A1:F1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left")

    if subtitle:
        ws.merge_cells("A2:F2")
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, italic=True, color="666666")
        header_row = 3
    else:
        header_row = 2

    header_fill = PatternFill(start_color="4F6FE8", end_color="4F6FE8", fill_type="solid")
    for col_idx, (_, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, start=header_row + 1):
        for col_idx, (key, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_cell_value(row.get(key, "")))
            cell.alignment = Alignment(vertical="center")

    widths = [18, 12, 28, 14, 14, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    last_row = ws.max_row
    ws.auto_filter.ref = f"A{header_row}:F{max(last_row, header_row)}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_pdf(rows: Iterable[Mapping], title: str = "Attendance Register", subtitle: str = "") -> BytesIO:
    """Build an in-memory landscape A4 PDF attendance register."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        title=title,
        author="Student Attendance System",
    )

    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"])]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 10))

    header = [label for _, label in COLUMNS]
    data = [header]
    for row in rows:
        data.append([_cell_value(row.get(key, "")) for key, _ in COLUMNS])

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F6FE8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E7EBF2")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FB")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf
